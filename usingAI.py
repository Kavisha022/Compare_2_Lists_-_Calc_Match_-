import streamlit as st
import pandas as pd
from sentence_transformers import SentenceTransformer, util
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO

# Load the SentenceTransformer model
@st.cache_resource
def load_model():
    return SentenceTransformer('all-MiniLM-L6-v2')

model = load_model()

st.title("📊 Semantic Word Matching Tool")
st.markdown("Upload an Excel file with **Sheet1** and **Sheet2**. The tool will match the **first column** from each sheet based on **semantic similarity** using a Hugging Face model.")

uploaded_file = st.file_uploader("📤 Upload Excel File", type=['xlsx'])

if uploaded_file:
    try:
        # Read the two sheets
        sheet1_df = pd.read_excel(uploaded_file, sheet_name='Sheet1')
        sheet2_df = pd.read_excel(uploaded_file, sheet_name='Sheet2')

        # Get the first column from both sheets
        list1 = sheet1_df.iloc[:, 0].dropna().astype(str).tolist()
        list2 = sheet2_df.iloc[:, 0].dropna().astype(str).tolist()

        st.info("🔍 Matching words...")

        # Generate embeddings for both lists
        emb1 = model.encode(list1, convert_to_tensor=True)
        emb2 = model.encode(list2, convert_to_tensor=True)

        matches = []

        # Set the match threshold (cosine similarity)
        match_threshold = 0.50  # Only consider matches above 50% similarity

        # Compare each word in Sheet1 with all words in Sheet2
        for word1, emb_word1 in zip(list1, emb1):
            cosine_scores = util.pytorch_cos_sim(emb_word1, emb2)  # Compare current word from Sheet1 to all words in Sheet2
            
            for idx, score in enumerate(cosine_scores[0]):
                # We need to compare word1 (from Sheet1) to word2 (from Sheet2)
                word2 = list2[idx]
                score_val = score.item()

                # Only consider matches above the threshold
                if score_val >= match_threshold:
                    matches.append((word1, word2, round(score_val * 100, 2)))

        result_df = pd.DataFrame(matches, columns=['Sheet1 Value', 'Best Match from Sheet2', 'Match Percentage'])
        st.success("✅ Matching complete!")
        st.dataframe(result_df)

        # Save to Excel
        output = BytesIO()
        result_df.to_excel(output, index=False)
        output.seek(0)

        # Load workbook for coloring
        wb = load_workbook(output)
        ws = wb.active

        # Define colors for formatting
        green_fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")   # Bright green
        blue_fill = PatternFill(start_color="FF66CCFF", end_color="FF66CCFF", fill_type="solid")    # Sky blue
        pink_fill = PatternFill(start_color="FFFFC0CB", end_color="FFFFC0CB", fill_type="solid")    # Light pink

        # Apply color formatting based on match percentage
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=3):
            score = row[2].value
            if score is None:
                continue
            if score >= 85:
                fill = green_fill
            elif score >= 60:
                fill = blue_fill
            else:
                fill = pink_fill
            for cell in row:
                cell.fill = fill

        # Save final colored Excel
        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        st.download_button("📥 Download Colored Excel", data=final_output, file_name="semantic_matches_sheetwise.xlsx")

    except Exception as e:
        st.error(f"❌ Error processing file: {e}")
