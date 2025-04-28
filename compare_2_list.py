import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from rapidfuzz import fuzz, distance

# Paths
file_path = r'D:\OneDrive - Lowcode Minds Technology Pvt Ltd\Desktop\Compare two list\dataset.xlsx'
output_path = r'D:\OneDrive - Lowcode Minds Technology Pvt Ltd\Desktop\Compare two list\matches_output.xlsx'

# Read data
sheet1 = pd.read_excel(file_path, sheet_name='Sheet1')
sheet2 = pd.read_excel(file_path, sheet_name='Sheet2')

# Extract lists
list1 = sheet1.iloc[:, 0].dropna().tolist()
list2 = sheet2.iloc[:, 0].dropna().tolist()

matches = []

# Set maximum edit distance allowed
max_distance = 2

# Compare every word in Sheet1 with Sheet2
for word1 in list1:
    for word2 in list2:
        # Calculate Levenshtein Distance (edit distance)
        edit_dist = distance.Levenshtein.distance(str(word1), str(word2))
        max_len = max(len(word1), len(word2))
        # Calculate similarity percentage
        similarity = (1 - edit_dist / max_len) * 100

        # If edit distance is small enough, consider it a match
        if edit_dist <= max_distance:
            matches.append((word1, word2, round(similarity, 2)))

# Save matches
result_df = pd.DataFrame(matches, columns=['Word from Sheet1', 'Similar Word from Sheet2', 'Match Percentage'])
result_df.to_excel(output_path, index=False)

# Load workbook for coloring
wb = load_workbook(output_path)
ws = wb.active

# Define fills
green_fill = PatternFill(start_color="FFFF69B4", end_color="FFFF69B4", fill_type="solid")  # Pink
blue_fill = PatternFill(start_color="FF87CEEB", end_color="FF87CEEB", fill_type="solid")   # Blue
yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid") # Yellow

# Apply color formatting based on match percentage
for row in ws.iter_rows(min_row=2, min_col=1, max_col=3):
    score = row[2].value
    if score is None:
        continue

    if score >= 85:
        fill = green_fill
    elif score >= 60:
        fill = blue_fill
    else:
        fill = yellow_fill

    for cell in row:
        cell.fill = fill

wb.save(output_path)
print(f"✅ Strict word-to-word matching complete! File saved to:\n{output_path}")
