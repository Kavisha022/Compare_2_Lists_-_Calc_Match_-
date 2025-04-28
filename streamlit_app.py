import streamlit as st
import pandas as pd
from rapidfuzz import distance
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO

st.title("📊 Strict Word Similarity Tool (Sheet1 vs Sheet2)")

uploaded_file = st.file_uploader("Upload Excel File with Sheet1 and Sheet2", type=['xlsx'])

if uploaded_file:
    try:
        # Load both sheets
        sheet1_df = pd.read_excel(uploaded_file, sheet_name='Sheet1')
        sheet2_df = pd.read_excel(uploaded_file, sheet_name='Sheet2')

        # Extract the first column from each sheet
        list1 = sheet1_df.iloc[:, 0].dropna().tolist()
        list2 = sheet2_df.iloc[:, 0].dropna().tolist()

        matches = []

        # Set minimum match percentage and maximum Levenshtein distance allowed for a match
        min_match_percentage = 50
        max_distance = 2

        # Compare each word in Sheet1 with all words in Sheet2
        for item1 in list1:
            for item2 in list2:
                # Calculate Levenshtein distance between item1 and item2
                dist = distance.Levenshtein.distance(str(item1), str(item2))
                max_len = max(len(item1), len(item2))
                similarity = (1 - dist / max_len) * 100

                # If Levenshtein distance is small enough and similarity is above the threshold, consider as a match
                if dist <= max_distance and similarity >= min_match_percentage:
                    matches.append((item1, item2, round(similarity, 2)))

        result_df = pd.DataFrame(matches, columns=['Word from Sheet1', 'Best Match from Sheet2', 'Match Percentage'])

        st.success("✅ Matching Completed!")
        st.dataframe(result_df)

        # Save the results to Excel
        output = BytesIO()
        result_df.to_excel(output, index=False)
        output.seek(0)

        # Load workbook and apply coloring
        wb = load_workbook(output)
        ws = wb.active

        # Define color fills
        green_fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")   # Green
        blue_fill = PatternFill(start_color="FFADD8E6", end_color="FFADD8E6", fill_type="solid")     # Light Blue
        pink_fill = PatternFill(start_color="FFFFC0CB", end_color="FFFFC0CB", fill_type="solid")     # Pink

        # Apply color formatting based on match score
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=3):
            score = row[2].value
            if score is None:
                continue
            if score >= 85:
                fill = green_fill
            elif score >= 60:
                fill = blue_fill
            else:
                fill = pink_fill
            for cell in row:
                cell.fill = fill

        # Save the colored workbook to BytesIO for download
        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        # Download button
        st.download_button("📥 Download Colored Excel", data=final_output, file_name="matches_output.xlsx")

    except Exception as e:
        st.error(f"❌ Error reading file or sheets: {e}")
